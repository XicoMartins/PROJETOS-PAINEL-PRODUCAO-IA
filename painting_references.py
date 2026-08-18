from __future__ import annotations

import json
import math
import re
import unicodedata
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from painting_colors import COLOR_BY_ABBREVIATION

ReferenceKey = tuple[str, str, str]


@dataclass(frozen=True)
class ReferenceFile:
    name: str
    size: int
    modified_ns: int


@dataclass(frozen=True)
class ReferenceSnapshot:
    directory: str
    files: tuple[ReferenceFile, ...]
    warnings: tuple[str, ...] = ()


@dataclass
class ReferenceCatalog:
    quantities: dict[ReferenceKey, float]
    warnings: tuple[str, ...] = ()


def normalize_reference_text(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    return " ".join(
        "".join(char for char in text if unicodedata.category(char) != "Mn")
        .strip()
        .upper()
        .split()
    )


def normalized_process(value: object) -> str:
    label = re.sub(r"\b(?:ENVIO|REMESSA|RETORNO)\b", " ", normalize_reference_text(value))
    label = re.sub(r"[-–—]+", " ", label)
    label = " ".join(label.split())
    for color in sorted(set(COLOR_BY_ABBREVIATION.values()), key=len, reverse=True):
        normalized_color = normalize_reference_text(color)
        if label == normalized_color:
            return "PROCESSO NAO INFORMADO"
        if label.endswith(f" {normalized_color}"):
            return label[: -len(normalized_color)].strip()
    return label or "PROCESSO NAO INFORMADO"


def reference_key(display: object, process: object, movement: str) -> ReferenceKey:
    return normalize_reference_text(display), normalized_process(process), movement


def complete_sets(total: float, qnt: float | None) -> int | None:
    if qnt is None or not math.isfinite(qnt) or qnt <= 0:
        return None
    return math.floor(max(0.0, float(total)) / qnt)


def movement_from_reference(tooling: object, process: object) -> str | None:
    for value in (tooling, process):
        text = normalize_reference_text(value)
        if "RETORNO" in text:
            return "retorno"
        if "ENVIO" in text or "REMESSA" in text:
            return "remessa"
    return None


def scan_reference_directory(directory: str | Path) -> ReferenceSnapshot:
    root = Path(directory)
    try:
        files = tuple(
            ReferenceFile(path.name, path.stat().st_size, path.stat().st_mtime_ns)
            for path in sorted(root.iterdir(), key=lambda item: item.name.casefold())
            if path.is_file()
            and path.suffix.casefold() == ".xlsx"
            and not path.name.startswith("~$")
        )
        return ReferenceSnapshot(str(root), files)
    except OSError as exc:
        return ReferenceSnapshot(str(root), (), (f"Pasta de referências indisponível: {exc}",))


def _collect_workbook_references(
    workbook: object,
    source_name: str,
    collected: dict[ReferenceKey, set[float]],
    warnings: list[str],
) -> None:
    sheet = workbook.active
    header_row = None
    indexes: dict[str, int] = {}
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1
    ):
        candidate = {
            normalize_reference_text(value): index
            for index, value in enumerate(row)
            if normalize_reference_text(value)
        }
        if {"ACABADO", "FERRAMENTAL", "PROCESSO", "QNT"} <= candidate.keys():
            header_row, indexes = row_number, candidate
            break
    if header_row is None:
        warnings.append(f"{source_name}: cabeçalhos ACABADO/FERRAMENTAL/PROCESSO/QNT ausentes")
        return
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        display = row[indexes["ACABADO"]] if indexes["ACABADO"] < len(row) else None
        tooling = row[indexes["FERRAMENTAL"]] if indexes["FERRAMENTAL"] < len(row) else None
        process = row[indexes["PROCESSO"]] if indexes["PROCESSO"] < len(row) else None
        raw_qnt = row[indexes["QNT"]] if indexes["QNT"] < len(row) else None
        if not any((display, tooling, process, raw_qnt)):
            continue
        movement = movement_from_reference(tooling, process)
        try:
            qnt = float(str(raw_qnt).replace(",", "."))
        except (TypeError, ValueError):
            qnt = 0
        if not display or not process or movement is None or not math.isfinite(qnt) or qnt <= 0:
            warnings.append(f"{source_name}: linha {row_number} sem referência QNT válida")
            continue
        collected.setdefault(reference_key(display, process, movement), set()).add(qnt)


def _catalog_from_collected(
    collected: dict[ReferenceKey, set[float]], warnings: list[str]
) -> ReferenceCatalog:
    quantities: dict[ReferenceKey, float] = {}
    for key, values in sorted(collected.items()):
        if len(values) == 1:
            quantities[key] = next(iter(values))
        else:
            warnings.append(f"Referência ambígua para {' / '.join(key)}: {sorted(values)}")
    return ReferenceCatalog(quantities, tuple(dict.fromkeys(warnings)))


def load_reference_catalog(snapshot: ReferenceSnapshot) -> ReferenceCatalog:
    collected: dict[ReferenceKey, set[float]] = {}
    warnings = list(snapshot.warnings)
    root = Path(snapshot.directory)
    for source in snapshot.files:
        workbook = None
        try:
            workbook = load_workbook(root / source.name, read_only=True, data_only=True)
            _collect_workbook_references(workbook, source.name, collected, warnings)
        except Exception as exc:
            warnings.append(f"{source.name}: não foi possível ler a planilha ({exc})")
        finally:
            if workbook is not None:
                workbook.close()
    return _catalog_from_collected(collected, warnings)


def _read_url(url: str, opener=urlopen) -> bytes:
    request = Request(
        quote(url, safe=":/?&=%+"),
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": "painel-pintura-mtech",
        },
    )
    with opener(request, timeout=12) as response:
        return response.read()


def load_github_reference_catalog(contents_url: str, opener=urlopen) -> ReferenceCatalog:
    collected: dict[ReferenceKey, set[float]] = {}
    warnings: list[str] = []
    try:
        listing = json.loads(_read_url(contents_url, opener).decode("utf-8"))
        if not isinstance(listing, list):
            raise ValueError("resposta de diretório inválida")
    except Exception as exc:
        return ReferenceCatalog({}, (f"Referências online indisponíveis: {exc}",))

    sources = sorted(
        (
            item
            for item in listing
            if isinstance(item, dict)
            and item.get("type") == "file"
            and str(item.get("name", "")).casefold().endswith(".xlsx")
            and not str(item.get("name", "")).startswith("~$")
        ),
        key=lambda item: str(item.get("name", "")).casefold(),
    )
    for source in sources:
        source_name = str(source.get("name", ""))
        download_url = str(source.get("download_url") or "")
        workbook = None
        try:
            if not download_url:
                raise ValueError("URL de download ausente")
            workbook = load_workbook(
                BytesIO(_read_url(download_url, opener)),
                read_only=True,
                data_only=True,
            )
            _collect_workbook_references(workbook, source_name, collected, warnings)
        except Exception as exc:
            warnings.append(f"{source_name}: não foi possível ler a planilha online ({exc})")
        finally:
            if workbook is not None:
                workbook.close()
    return _catalog_from_collected(collected, warnings)
